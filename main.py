from bs4 import BeautifulSoup as bs
import requests
import yagmail
from openpyxl import load_workbook

wb = load_workbook("amazon_tracker.xlsx")
ws = wb.active
row_count = ws.max_row

print("Processing...........")

def send_alert(reciever):

    for col in range(2, row_count+1):
        link = ws[f"B{col}"].value
        initial_price = ws[f"C{col}"].value
        name = ws[f"A{col}"].value.lstrip()

        response = requests.get(link, headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36", "Accept-Language": "en-US,en;q=0.9"})
        html = response.text
        data = bs(html, "html.parser")
        latest_price = data.find("span", class_="a-price-whole")
        latest_price = float(latest_price.get_text().replace(",", ''))



        if (latest_price < initial_price):
            subject = f"Price Drop Alert: {name}"
            contents = [
                f"We are pleased to inform you that the price of '{name}' on Amazon has dropped to a new low of Rs,{latest_price}. This price reduction is below your desired threshold of Rs{initial_price}, making it an excellent time to consider your purchase.\n\n",

                f"Product Details:\n\n",
                f"Product Name: {name}\n\n",
                f"Current Price: Rs,{latest_price}\n",
                f"Original Price: Rs,{initial_price}\n\n",
                f"Product Link: {link}\n\n",

                f"Hurry and seize this opportunity to make your purchase while the price is favorable.\n\n",

                f"Thank you for using our Amazon price tracking service",
            
            ]
            contents = ''.join(contents)

            yagmail.SMTP('Email Username', 'Email Password').send(reciever, subject, contents)
            print(f"Price Drop Detected! . Succesfully Send The Alert Message To {reciever}")

        ws[f"C{col}"].value = latest_price




if __name__ == "__main__":

    reciever = "ashharjosh@gmail.com" # Destination Mail For Alert Messages

    try:
        send_alert(reciever)
    except:
        print("Error Tracking The Prices")

    wb.save("amazon_tracker.xlsx")
    print("Finished Checking For Price Alert")
